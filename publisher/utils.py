# -*- coding: utf-8 -*-
# 功能函数放置在此文件中
# from forms import LoginForm, RegisterForm
from django.db import models
from users import models
import datetime
from openpyxl.reader.excel import load_workbook
import os


def add_publisher(form):
    data = form.cleaned_data
    publisher = models.Adminpublisher.objects.create(loginname=data['loginname'],
                                                     password=data['password'],
                                                     telephone=data['telephone'],
                                                     email=data['email'],
                                                     createtime=datetime.datetime.now(),
                                                     id_hospital_id=data['hospital'])
    publisher.save()


def auth_publisher(form):
    if form.is_valid():
        data = form.cleaned_data
        loginname = data['loginname']
        password = data['password']
        # check
        publisher = models.Adminpublisher.objects.filter(loginname=loginname, password=password)
        # print publisher
        return publisher
    else:
        return None


def select_hospital():
    hospital = models.Hospital.objects.all().values('id_hospital', 'name')
    h_list=[]
    if hospital:
        for h_dict in hospital:
            t_tuple = (h_dict['id_hospital'], h_dict['name'])
            h_list.append(t_tuple)
    return h_list


def publisher_exists(name):
    publisher = models.Adminpublisher.objects.filter(loginname=name)
    if publisher:
        return True
    return False


def select_all_bulletin(hospital_id):
    bulletins = models.Bulletin.objects.filter(id_adminpublisher__id_hospital_id=hospital_id)
    return bulletins


def get_publisher_info(_loginname):
    '''
    :param _loginname:
    :return: id_hospital and id_adminpubisher given pulisher's logining name
    '''
    ids = models.Adminpublisher.objects.filter(loginname=_loginname).values('id_hospital_id', 'id_adminpublisher')
    id_hospital = 0
    id_adminpublisher = 0
    for id in ids:
        id_hospital = id['id_hospital_id']
        id_adminpublisher = id['id_adminpublisher']
    return (id_hospital,id_adminpublisher)


def get_doctor_info(_doctor_id):
    '''
    :param _doctor_id:
    :return: id_doctor_department given doctor_id
    '''
    id_doctor_department = 0
    ids = models.DoctorDepartment.objects.filter(id_doctor_id=_doctor_id).\
        values('id_doctor_department')
    for id in ids:
        id_doctor_department = id['id_doctor_department']
    return id_doctor_department


def getDepartments(hospital_id):
    apartment_list = models.Department.objects.filter(id_hospital_id=hospital_id).\
        values_list('id_department', 'name')
    return list(apartment_list)


def getDoctors(hospital_id):
    doctor_list = models.DoctorDepartment.objects.filter(id_department__id_hospital_id=hospital_id).\
        values_list('id_doctor_id', 'id_doctor__name')
    return list(doctor_list)


def create_bulletin(form,session):
    data = form.cleaned_data
    appointment = models.Bulletin.objects.create(createtime=datetime.datetime.now(),
                                                 availabletime=data['available_time'],
                                                 fee=data['fee'],
                                                 countavailable=data['count_available'],
                                                 countoccupied=data['count_occupied'],
                                                 id_adminpublisher_id=session.get('publisher_id'),
                                                 id_doctor_department_id=get_doctor_info(data['doctor']))
    appointment.save()


def alter_bulletin(form,session):
    data = form.cleaned_data
    models.Bulletin.objects.filter(id_bulletin=session['id_bulletin']).update\
        (createtime=datetime.datetime.now(),
         availabletime=data['available_time'],
         fee=data['fee'],
         countavailable=data['count_available'],
         countoccupied=data['count_occupied'],
         id_adminpublisher_id=session.get('publisher_id'),
         id_doctor_department_id=get_doctor_info(data['doctor']))


def handle_uploaded_file(request, f=None):
    datenow = datetime.datetime.now()
    filedate = datenow.strftime('%Y%m%d-%H%M%S')
    path = os.path.join(os.path.abspath('.'),'publisher','fileUpload')
    filepath = path + '/' + filedate + '_' + f.name
    with open(filepath, 'ab') as de:
        for chunk in f.chunks():
            de.write(chunk)
    wb = load_workbook(filepath)
    print(filepath)
    table = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    bulletin_list = []
    for i in range(2, table.max_row + 1):
        if table.cell(row=i, column=1).value is None:
            # print '科室为空，应跳过'
            continue
        print('正在导入第' + str(i - 1) + '行...')

        _id_doctor_department = _get_id_doctor_department(request, dept_name=table.cell(row=i, column=1).\
                                                          value, dt_name=table.cell(row=i, column=2).value)
        _availabletime = table.cell(row=i, column=3).value
        #数据库已存在相同的预约发布信息，应跳过
        if isItemRepeated(_id_doctor_department,_availabletime):
            continue
        bulletin = models.Bulletin(
            availabletime=_availabletime,
            countavailable=table.cell(row=i, column=4).value,
            countoccupied=table.cell(row=i, column=5).value,
            fee=table.cell(row=i, column=6).value,
            createtime=datetime.datetime.now(),
            id_adminpublisher_id=request.session['publisher_id'],
            id_doctor_department_id=_id_doctor_department
        )
        bulletin_list.append(bulletin)
    models.Bulletin.objects.bulk_create(bulletin_list)


def _get_id_doctor_department(request, dept_name, dt_name):
    import sys
    reload(sys)
    sys.setdefaultencoding('utf-8')
    doc_dept = models.DoctorDepartment.objects.get(id_department__name=dept_name,
                                                   id_department__id_hospital_id=request.session['hospital_id'],
                                                   id_doctor__name=dt_name)
    return doc_dept.id_doctor_department


def isItemRepeated(_id_doctor_department,_availabletime):
    all_availabetime = models.Bulletin.objects.filter(id_doctor_department=_id_doctor_department).values('availabletime')
    for availabletime in all_availabetime:
        availabletime = str(availabletime['availabletime']+datetime.timedelta(hours=8))[0:-6]
        if availabletime == str(_availabletime):
            return True
    return False